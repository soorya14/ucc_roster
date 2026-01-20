
#!/usr/bin/env python3
"""
Build GCC-style monthly roster from 'employees (1).csv' (leave tracker style).
 
- Title row (row 1) contains date labels like: 01-Jan, 02-Jan, ...
- Header row (row 2) contains: SL.No, Associate Name, Gender, Reporting to, Support , Location, <DoW...>
- Data rows: one per associate; any non-empty cell under date columns is a code that overrides the computed shift.
 
Shift rules supported (Mon–Fri only; weekends = WO):
  1) G-only (names list)
  2) B↔G weekly rotation (names list) — defaults to start on B in week 1 of the month
  3) Everyone else → ABC weekly rotation
 
Leaves/holidays override shifts: Leave, HO, CO-HO, AD-Leave, WO, etc.
 
Usage examples:
  # January 2026, default lists hard-coded in this file:
  python build_Roster_CreateExcel.py --csv "employees (1).csv" --year 2026 --month 1 --output "GCC_Roster_Jan2026.xlsx"
 
  # Provide lists inline (comma-separated), if you want to override in CLI:
  python build_Roster_CreateExcel.py --csv "employees (1).csv" --year 2026 --month 1 \
    --g-only "Manikandan S,Sindhu Ramesh,Ramanan K A V,Sudarshan Iyengar,Praveen AV" \
    --bg-rotate "Naveen Kumar,Kumaragurunaathan,G Monica,Angelin S,Suruthinathan" \
    --output "GCC_Roster_Jan2026.xlsx"
"""
 
import argparse
import csv
from datetime import datetime, date, timedelta
from pathlib import Path
from time import sleep
 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
 
 
# -------------------------
# Look & Feel (colors, fonts)
# -------------------------
 
FILLS = {
    "A": PatternFill("solid", fgColor="CCFFCC"),   # light green
    "B": PatternFill("solid", fgColor="CCE5FF"),   # light blue
    "C": PatternFill("solid", fgColor="FFCCCC"),   # light red
    "G": PatternFill("solid", fgColor="FFD966"),   # amber
    "WO": PatternFill("solid", fgColor="FFF2CC"),  # pale yellow
    "Leave": PatternFill("solid", fgColor="D9D9D9"),
    "HO": PatternFill("solid", fgColor="E2EFDA"),      # holiday
    "CO-HO": PatternFill("solid", fgColor="C9C9C9"),   # comp-off holiday
    "AD-Leave": PatternFill("solid", fgColor="F4B084"), # additional/admin leave
}
 
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT  = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")
 
 
# -------------------------
# Helpers
# -------------------------
 
def month_days(year: int, month: int):
    """Return list[date] for all days in month."""
    start = date(year, month, 1)
    stop = date(year + (month == 12), (month % 12) + 1, 1) - timedelta(days=1)
    cur = start
    out = []
    while cur <= stop:
        out.append(cur)
        cur += timedelta(days=1)
    return out
 
def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())
 
def norm_code(x: str) -> str:
    """Normalize textual codes to canonical forms used in output."""
    if not x:
        return ""
    s = x.strip()
    if not s:
        return ""
    s = s.replace(" ", "").upper()
    # special variants
    if s in {"CO-HO", "COHO", "CO- HO", "CO -HO".replace(" ", ""), "CO - HO".replace(" ", "")}:
        return "CO-HO"
    if s in {"AD-LEAVE", "ADLEAVE", "AD- LEAVE".replace(" ", "")}:
        return "AD-Leave"
    if s in {"WEEKOFF", "WEEK-OFF", "OFF"}:
        return "WO"
    if s in {"HOLIDAY"}:
        return "HO"
    # standard ones
    if s in {"A","B","C","G","WO","HO","LEAVE"}:
        return "Leave" if s == "LEAVE" else s
    # pass through any custom code (e.g., SL/CL/PL if you later add)
    return s
 
def parse_name_list(arg: str):
    """Return a set of names from a comma-separated argument (or empty if None)."""
    if not arg:
        return set()
    return {p.strip() for p in arg.split(",") if p.strip()}
 
 
# -------------------------
# CSV Loader (your "employees (1).csv" format)
# -------------------------
 
def load_employees_and_leaves(csv_path: Path, year: int, month: int):
    """
    Reads the 'employees (1).csv' structure:
      - title row (row 1) contains date labels like '01-Jan, 02-Jan, ...'
      - header row (row 2) contains the static columns + DoW headings
      - subsequent rows are associates
 
    Returns:
      employees: list of dicts with keys:
        sl, name, gender, reporting_to, support, location, leaves(dict[date]=code)
      issues: list[(level, msg)]
    """
    issues = []
    with csv_path.open(newline="", encoding="utf-8") as f:
        rdr = csv.reader(f)
        rows = list(rdr)
 
    if len(rows) < 3:
        raise ValueError("CSV missing rows (need title row + header row + data rows)")
 
    title_row  = rows[0]
    header_row = rows[1]
    data_rows  = rows[2:]
 
    # Build date columns mapping from title row (index>=6)
    date_cols = {}  # idx -> date
    for i, lbl in enumerate(title_row[6:], start=6):
        lbl = (lbl or "").strip()
        if not lbl:
            continue
        # examples: 01-Jan, 31-Jan
        try:
            d = datetime.strptime(f"{lbl}-{year}", "%d-%b-%Y").date()
        except Exception:
            d = None
        if d and d.month == month:
            date_cols[i] = d
 
    if not date_cols:
        issues.append(("ERROR", f"No date columns in title row matched {month:02d}-{year}"))
        # Continue; we'll still read static data.
 
    # Build employees
    employees = []
    seen_names = set()
    for rnum, r in enumerate(data_rows, start=3):
        if not any(r):
            continue
        # Ensure we have enough columns to index safely
        need_cols = max(date_cols.keys(), default=5)
        if len(r) <= need_cols:
            r = list(r) + [""] * (need_cols + 1 - len(r))
 
        sl       = (r[0] or "").strip()
        name     = (r[1] or "").strip()
        gender   = (r[2] or "").strip()
        manager  = (r[3] or "").strip()
        support  = (r[4] or "").strip()
        location = (r[5] or "").strip()
        if not name:
            issues.append(("WARNING", f"Row {rnum}: empty Associate Name -> skipped"))
            continue
        if name in seen_names:
            issues.append(("WARNING", f"Duplicate associate name '{name}' (row {rnum})"))
        seen_names.add(name)
 
        leaves = {}
        for ci, d in date_cols.items():
            code = norm_code(r[ci])
            if code:
                leaves[d] = code
 
        employees.append({
            "sl": sl,
            "name": name,
            "gender": gender,
            "reporting_to": manager,
            "support": support,
            "location": location,
            "leaves": leaves,
        })
 
    if not employees:
        issues.append(("ERROR", "No employees parsed from CSV"))
    return employees, issues
 
 
# -------------------------
# Shift Engine
# -------------------------
 
def build_month_grid(employees, year: int, month: int,
                     g_only: set, bg_rotate: set,
                     start_bg: str = "B", abc_order=("A","B","C")):
    """
    Returns:
      days: list[date]
      grid: list[list[str]] -> per-employee, per-day code
      issues: list[(level, msg)]
    """
    issues = []
    days   = month_days(year, month)
    base_monday = monday_of_week(days[0])
 
    # sanity on start_bg
    start_bg = start_bg.upper()
    if start_bg not in {"B","G"}:
        start_bg = "B"
        issues.append(("INFO", "start_bg not in {B,G}; defaulted to B"))
 
    emp_names = {e["name"] for e in employees}
    for s_name in sorted(g_only - emp_names):
        issues.append(("WARNING", f"G-only name not in CSV: '{s_name}'"))
    for s_name in sorted(bg_rotate - emp_names):
        issues.append(("WARNING", f"B/G-rotate name not in CSV: '{s_name}'"))
 
    grid = []
    for emp in employees:
        row_codes = []
        for d in days:
            # weekend rule
            if d.weekday() >= 5:
                base = "WO"
            else:
                if emp["name"] in g_only:
                    base = "G"
                elif emp["name"] in bg_rotate:
                    rel_week = (monday_of_week(d) - base_monday).days // 7
                    # alternate starting with start_bg
                    if start_bg == "B":
                        base = "B" if (rel_week % 2 == 0) else "G"
                    else:
                        base = "G" if (rel_week % 2 == 0) else "B"
                else:
                    # ABC weekly rotation
                    rel_week = (monday_of_week(d) - base_monday).days // 7
                    base = abc_order[rel_week % 3]
            # leaves override
            code = emp["leaves"].get(d, base)
            row_codes.append(code)
        grid.append(row_codes)
    return days, grid, issues
 
 
# -------------------------
# Workbook Builder
# -------------------------
 
def build_workbook(employees, days, grid, year, month):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{date(year, month, 1):%b %Y}"
 
    # Title row
    total_cols = 6 + len(days)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1, value="GCC MONTHLY ROSTER")
    t.font = TITLE_FONT
    t.alignment = CENTER
 
    # Headers
    headers = ["SL.No", "Associate Name", "Gender", "Reporting to", "Support ", "Location"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.alignment = CENTER
    for j, d in enumerate(days, start=1):
        c = ws.cell(row=2, column=6 + j, value=d.strftime("%d/%m/%Y"))
        c.font = HEADER_FONT
        c.alignment = CENTER
 
    # Data rows
    r0 = 3
    for r_idx, (emp, codes) in enumerate(zip(employees, grid), start=r0):
        ws.cell(row=r_idx, column=1, value=emp["sl"]).alignment = CENTER
        ws.cell(row=r_idx, column=2, value=emp["name"]).alignment = LEFT
        ws.cell(row=r_idx, column=3, value=emp["gender"]).alignment = CENTER
        ws.cell(row=r_idx, column=4, value=emp["reporting_to"]).alignment = LEFT
        ws.cell(row=r_idx, column=5, value=emp["support"]).alignment = LEFT
        ws.cell(row=r_idx, column=6, value=emp["location"]).alignment = LEFT
        for j, code in enumerate(codes, start=1):
            c = ws.cell(row=r_idx, column=6 + j, value=code)
            c.alignment = CENTER
            if code in FILLS:
                c.fill = FILLS[code]
 
    # Widths & freeze panes
    widths = [8, 28, 8, 20, 30, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in range(7, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 5
    ws.freeze_panes = ws.cell(row=3, column=7)
 
    # Totals block (per-day counts)
    base_row = r0 + len(employees) + 1
    codes_for_totals = ["G", "A", "B", "C", "WO", "Leave", "HO", "CO-HO", "AD-Leave"]
    for idx, code in enumerate(codes_for_totals):
        r = base_row + idx
        ws.cell(row=r, column=1, value=f"{code} Count").font = HEADER_FONT
        for j in range(len(days)):
            cnt = sum(1 for row in grid if row[j] == code)
            c = ws.cell(row=r, column=6 + j + 1, value=cnt)
            c.alignment = CENTER
 
    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend.append(["Code", "Meaning"])
    legend["A1"].font = legend["B1"].font = HEADER_FONT
    for code, meaning in [
        ("A", "Shift A (06:30–15:30 IST)"),
        ("B", "Shift B (14:00–23:00 IST)"),
        ("C", "Shift C (22:00–07:00 IST)"),
        ("G", "General (11:00–20:00 IST)"),
        ("WO", "Weekly Off (Saturday, Sunday)"),
        ("HO", "Holiday"),
        ("CO-HO", "Comp-Off on Holiday"),
        ("Leave", "Leave (generic, from CSV)"),
        ("AD-Leave", "Additional/Admin Leave"),
    ]:
        legend.append([code, meaning])
 
    return wb
 
 
def attach_validation_sheet(wb: Workbook, issues):
    ws = wb.create_sheet("Validation")
    ws.append(["Level", "Message"])
    ws["A1"].font = ws["B1"].font = HEADER_FONT
    for level, msg in issues:
        ws.append([level, msg])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 120
 
 
def save_workbook_safely(wb: Workbook, output_file: Path, retries=3, delay=1.5) -> Path:
    """Save with retries and timestamp fallback to avoid Windows/OneDrive lock errors."""
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            wb.save(output_file)
            return output_file
        except PermissionError as e:
            last_err = e
            print(f"⚠️ PermissionError saving '{output_file}' (attempt {attempt}/{retries}). Retrying...")
            sleep(delay)
    # Fallback
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    alt = output_file.with_name(f"{output_file.stem}_{ts}{output_file.suffix}")
    wb.save(alt)
    print(f"✅ Saved with fallback name: {alt}")
    return alt
 
 
# -------------------------
# Main
# -------------------------
 
def main():
    p = argparse.ArgumentParser(description="Create GCC-style roster from 'employees (1).csv'")
    p.add_argument("--csv", required=True, help="Path to 'employees (1).csv'")
    p.add_argument("--year", type=int, required=True, help="Target year (e.g., 2026)")
    p.add_argument("--month", type=int, required=True, help="Target month number (1-12)")
    p.add_argument("--output", default=None, help="Output Excel path (.xlsx)")
    p.add_argument("--g-only", default=None, help="Comma-separated names that are G-only")
    p.add_argument("--bg-rotate", default=None, help="Comma-separated names that rotate B↔G")
    p.add_argument("--start-bg", default="B", help="First week shift for B/G group: B or G (default B)")
    args = p.parse_args()
 
    csv_path = Path(args.csv)
    if not csv_path.exists():
        raise FileNotFoundError(csv_path)
 
    # Default lists (as per your last instruction)
    default_g_only = {
        "Manikandan S",
        "Sindhu Ramesh",
        "Ramanan K A V",
        "Sudarshan Iyengar",
        "Praveen AV",
    }
    default_bg_rotate = {
        "Naveen Kumar",
        "Kumaragurunaathan",
        "G Monica",
        "Angelin S",
        "Suruthinathan",
    }
 
    g_only   = parse_name_list(args.g_only)   or default_g_only
    bg_rot   = parse_name_list(args.bg_rotate) or default_bg_rotate
 
    # Load employees & leaves
    employees, issues = load_employees_and_leaves(csv_path, args.year, args.month)
 
    # Build grid
    days, grid, more_issues = build_month_grid(
        employees, args.year, args.month,
        g_only=g_only, bg_rotate=bg_rot,
        start_bg=args.start_bg.upper(), abc_order=("A","B","C")
    )
    issues.extend(more_issues)
 
    # Build workbook
    wb = build_workbook(employees, days, grid, args.year, args.month)
    attach_validation_sheet(wb, issues)
 
    # Default output name
    if args.output:
        out_path = Path(args.output)
    else:
        out_path = Path(f"GCC_Roster_{date(args.year, args.month, 1):%b%Y}.xlsx")
 
    saved = save_workbook_safely(wb, out_path)
    print(f"✅ Created: {saved}  (Employees: {len(employees)}, Days: {len(days)})")
 
if __name__ == "__main__":
    main()
 